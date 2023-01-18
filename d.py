import pandas as pd
from openpyxl import load_workbook

#file = r'C:\Users\Тагиров2\PycharmProjects\pythonProject_Exel\20_Raspisanie_16_01_2023-21_01_2023.xlsx'
file2= r"20_Raspisanie_16_01_2023-21_01_2023 (2).xlsx"
xl = pd.ExcelFile(file2)
print(xl.sheet_names)
#df1 = xl.parse('Лист1')
df1 = xl.parse('AllPages')
# Load in the workbook
wb = load_workbook(file2)

# Get sheet names
sheet=wb.get_sheet_by_name('AllPages')
print(sheet)

#поиск группы столбец j=103
day = []
for i in range(1,250):
  for j in range(1, 630):
     if(sheet.cell(row=i, column=j).value == "ИС-2"):
       print(i,j)
       IS2_column=j
       day.append(i)

#first day of the week IS-2
print("monday")
for i in range(day[1], day[2]):
  if (sheet.cell(row=i, column=IS2_column).value):
    #print(sheet.cell(row=i, column=1).value, end=" | ")  #день недели
    print(sheet.cell(row=i, column=2).value, end=" | ")#урок
    print(sheet.cell(row=i, column=3).value, end=" | ")#время
    print(sheet.cell(row=i, column=IS2_column).value, end=" | ")#predmet
    print(sheet.cell(row=i, column=IS2_column+1).value, end=" | ")#teacher
    print(sheet.cell(row=i, column=IS2_column+2).value)#cab

#first day of the week IS-2
print("suturday")
for i in range(day[5], 85):
  if (sheet.cell(row=i, column=IS2_column).value):
    #print(sheet.cell(row=i, column=1).value, end=" | ")  #день недели
    print(sheet.cell(row=i, column=2).value, end=" | ")#урок
    print(sheet.cell(row=i, column=3).value, end=" | ")#время
    print(sheet.cell(row=i, column=IS2_column).value, end=" | ")#predmet
    print(sheet.cell(row=i, column=IS2_column+1).value, end=" | ")#teacher
    print(sheet.cell(row=i, column=IS2_column+2).value)#cab